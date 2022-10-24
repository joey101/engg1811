import openpyxl


rankfcf = ["CPT", "FO", "SI", "CS", "FA"]
    
fleet = ["EAFC", "EACC", "SAFC", "SACC"]
full_fleet = "EAFC/SAFC/EACC/SACC"
fc = "EAFC/SAFC"
cc = "EACC/SACC"

wb = load_workbook(filename='E:\OneDrive - UNSW\ID\Qantas\ql.xlsx')
ws = wb.worksheets[0]
row_num = 2
for idx in range(len(worksheet.rows())):
    if row[idx].cell(row_num,4) == full_fleet:
        tmp = idx
        insert_rows(tmp + 1,8)
        for row in :
            i = rank[j]
            j++
         for i in 8 rows in col 5:
            i = fleet[j]
            j++ 
        row_num++
    elif row[idx].cell(row_num,4) = fc:
         tmp = idx
         insert_rows(tmp + 1,4)
          for i in 4 rows :col 4:
              
               i = rank[j]
               j++ only up to 2
         for i in 4 rows :col 5:
              
               i = rank[j]
               j++ only up to 2
        row_num++
    elif row[idx].cell(row_num,4) = cc:
         tmp = idx
         insert_rows(tmp + 1,4)
          for i in 4 rows :col 4:
              
               i = rank[j]
               j++ only up to 2
         for i in 4 rows :col 5:
              
               i = rank[j]
               j++ only up to 2
        row_num++
