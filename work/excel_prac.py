# -*- coding: utf-8 -*-
"""
Created on Mon Sep 26 21:47:49 2022

@author: Jawad
"""

from openpyxl import load_workbook


rankfcf = ["CPT", "FO", "SI", "CS", "FA"]
    
fleet = ["EAFC", "EACC", "SAFC", "SACC"]
full_fleet = "EAFC/SAFC/EACC/SACC"
fc = "EAFC/SAFC"
cc = "EACC/SACC"


wb = load_workbook(filename='E:\OneDrive - UNSW\ID\Qantas\ql.xlsx')
ws = wb.worksheets[0]
row_num = 2
for row in ws.iter_rows(min_row=2,min_col=1,max_row=95,max_col=9):
    for tmp,cell in enumerate(row): 
        if cell.value == full_fleet:
            ws.insert_rows(tmp + 1,8)
            print(tmp, cell.value)
